"""
Utility script to automatically run all problem instances with all algorithms
and save visualisations into their respective output folders.

Algorithms:
- Genetic Algorithm (GA)
- Random Search (RS)
- Greedy Search (GR)
"""

import os
import time
from typing import List, Dict

from main import (
    get_instance,
    list_instances,
    CargoVisualiser,
)
from genetic_cargo import GeneticAlgorithm
from random_cargo import RandomSearch
from greedy_cargo import GreedySearch


def run_and_save(instance_name: str, algo: str) -> Dict:
    """Run a single algorithm on a single instance and save visualisation.
    
    Returns:
        Dictionary with results data for Excel export
    """
    start_time = time.time()
    cargo_items, container = get_instance(instance_name)

    if algo == "GA":
        ga = GeneticAlgorithm(cargo_items, container)
        solution = ga.run(verbose=True)
        algo_abbrev = "GA"
    elif algo == "RS":
        rs = RandomSearch(cargo_items, container)
        solution = rs.run(max_iterations=2000, verbose=True)
        algo_abbrev = "RS"
    elif algo == "GR":
        gs = GreedySearch(cargo_items, container)
        solution = gs.run(verbose=True)
        algo_abbrev = "GR"
    else:
        raise ValueError(f"Unknown algorithm: {algo}")

    elapsed_time = time.time() - start_time

    # Prepare visualisation
    vis = CargoVisualiser(solution)
    vis.draw(title=instance_name)

    # Ensure algorithm-specific output directory exists (e.g. ./output/GA)
    base_output_dir = "./output"
    algo_output_dir = os.path.join(base_output_dir, algo_abbrev)
    os.makedirs(algo_output_dir, exist_ok=True)

    filename = os.path.join(
        algo_output_dir,
        f"{instance_name}_{algo_abbrev}_fitness_{solution.fitness:.2f}.png",
    )

    import matplotlib.pyplot as plt

    plt.savefig(filename, dpi=300, bbox_inches="tight", facecolor="#01364C")
    plt.close()

    print(f"[{algo_abbrev}] {instance_name}: visualisation saved to {filename}")

    # Collect data for Excel export
    num_placed = sum(1 for c in solution.cargo_items if c.placed)
    total_items = len(solution.cargo_items)
    com_x, com_y = solution.get_center_of_mass() if solution.complete else (0.0, 0.0)
    perfect = "Yes" if solution.fitness == 0.0 else "No"
    
    return {
        "Instance": instance_name,
        "Algorithm": algo_abbrev,
        "Fitness": solution.fitness,
        "Items Placed": num_placed,
        "Total Items": total_items,
        "Placement Rate (%)": round((num_placed / total_items) * 100, 2) if total_items > 0 else 0.0,
        "COM X": round(com_x, 3),
        "COM Y": round(com_y, 3),
        "Perfect Solution": perfect,
        "Time (seconds)": round(elapsed_time, 2),
        "Container Width": container.width,
        "Container Depth": container.depth,
        "Max Weight": container.max_weight,
        "Total Weight": round(sum(c.weight for c in solution.cargo_items if c.placed), 2),
    }


def export_to_excel(results: List[Dict], output_dir: str = "./output"):
    """Export results to an Excel spreadsheet."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("\nWarning: openpyxl not installed. Installing now...")
        import subprocess
        subprocess.check_call(["pip", "install", "openpyxl"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Headers
    headers = [
        "Instance",
        "Algorithm",
        "Fitness",
        "Items Placed",
        "Total Items",
        "Placement Rate (%)",
        "COM X",
        "COM Y",
        "Perfect Solution",
        "Time (seconds)",
        "Container Width",
        "Container Depth",
        "Max Weight",
        "Total Weight",
    ]
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_idx, result in enumerate(results, start=2):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=result.get(header, ""))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Highlight perfect solutions
            if header == "Perfect Solution" and result.get(header) == "Yes":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif header == "Fitness" and result.get(header) == 0.0:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for row_idx in range(2, len(results) + 2):
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value)
            if len(cell_value) > max_length:
                max_length = len(cell_value)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 20)

    # Save file
    os.makedirs(output_dir, exist_ok=True)
    excel_filename = os.path.join(output_dir, "results.xlsx")
    wb.save(excel_filename)
    print(f"\n{'=' * 70}")
    print(f"Excel file saved to: {excel_filename}")
    print(f"{'=' * 70}")


def run_all_instances():
    """Run all algorithms on all available instances."""
    instances = list_instances()
    algorithms = ["GA", "RS", "GR"]
    results = []

    print(f"Found {len(instances)} instances.")
    print(f"Running algorithms: {', '.join(algorithms)}")

    for name in instances:
        print("\n" + "=" * 70)
        print(f"INSTANCE: {name}")
        print("=" * 70)
        for algo in algorithms:
            print("\n" + "-" * 40)
            print(f"Running {algo} on {name}")
            print("-" * 40)
            result = run_and_save(name, algo)
            results.append(result)

    # Export to Excel
    export_to_excel(results)


if __name__ == "__main__":
    run_all_instances()

